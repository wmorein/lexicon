#!/usr/bin/env python3
"""
Excel file utilities using openpyxl.
"""

import argparse
import csv
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def get_sheet_names(xlsx_path: str) -> list[str]:
    """Return list of sheet names in workbook."""
    wb = load_workbook(xlsx_path, read_only=True)
    return wb.sheetnames


def read_sheet_data(xlsx_path: str, sheet_name: str | None = None) -> list[list]:
    """Read all data from a sheet as list of lists."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def to_csv(xlsx_path: str, csv_path: str, sheet_name: str | None = None):
    """Convert Excel sheet to CSV."""
    data = read_sheet_data(xlsx_path, sheet_name)
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(data)
    print(f"Written to {csv_path}")


def to_json(xlsx_path: str, json_path: str, sheet_name: str | None = None):
    """Convert Excel sheet to JSON (list of objects using first row as headers)."""
    data = read_sheet_data(xlsx_path, sheet_name)
    if not data:
        result = []
    else:
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(data[0])]
        result = [dict(zip(headers, row)) for row in data[1:]]

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, default=str)
    print(f"Written to {json_path}")


def summarize(xlsx_path: str):
    """Print summary of workbook structure."""
    wb = load_workbook(xlsx_path, read_only=True)
    print(f"Workbook: {xlsx_path}")
    print(f"Sheets: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  - {name}: {ws.max_row} rows x {ws.max_column} cols")


def main():
    parser = argparse.ArgumentParser(description="Excel file utilities")
    subparsers = parser.add_subparsers(dest='command', required=True)

    # sheets command
    p_sheets = subparsers.add_parser('sheets', help='List sheet names')
    p_sheets.add_argument('xlsx', help='Input Excel file')

    # to-csv command
    p_csv = subparsers.add_parser('to-csv', help='Convert to CSV')
    p_csv.add_argument('xlsx', help='Input Excel file')
    p_csv.add_argument('output', help='Output CSV file')
    p_csv.add_argument('--sheet', help='Sheet name (default: active sheet)')

    # to-json command
    p_json = subparsers.add_parser('to-json', help='Convert to JSON')
    p_json.add_argument('xlsx', help='Input Excel file')
    p_json.add_argument('output', help='Output JSON file')
    p_json.add_argument('--sheet', help='Sheet name (default: active sheet)')

    # summarize command
    p_sum = subparsers.add_parser('summarize', help='Summarize workbook')
    p_sum.add_argument('xlsx', help='Input Excel file')

    args = parser.parse_args()

    if args.command == 'sheets':
        for name in get_sheet_names(args.xlsx):
            print(name)
    elif args.command == 'to-csv':
        to_csv(args.xlsx, args.output, getattr(args, 'sheet', None))
    elif args.command == 'to-json':
        to_json(args.xlsx, args.output, getattr(args, 'sheet', None))
    elif args.command == 'summarize':
        summarize(args.xlsx)


if __name__ == '__main__':
    main()
